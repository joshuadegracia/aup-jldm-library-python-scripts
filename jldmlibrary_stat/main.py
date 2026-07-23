from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import mysql.connector as mariadb
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.widgets import DateEntry

# --- Database Connection Settings ---
DB_CONFIG = {
    "host": "192.168.2.104",
    "port": 3308,
    "user": "root",  # Update username if necessary
    "password": "@dm!N2026",  # Update password
    "database": "jldmlibrary_stat",
}


def fetch_sections():
    """Dynamically fetches active sections for the Area dropdown."""
    query = """
        SELECT id, name 
        FROM jldmlibrary_sections 
        WHERE is_active = 1 AND (is_deleted IS NULL OR is_deleted = 0)
        ORDER BY id ASC;
    """
    sections = {}
    try:
        conn = mariadb.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(query)
        for sec_id, name in cursor.fetchall():
            sections[name] = sec_id
        conn.close()
    except Exception as e:
        messagebox.showerror("Database Error", f"Failed to fetch sections:\n{e}")
    return sections


def fetch_date_bounds(section_id):
    """Pulls the MIN and MAX dates present in usage_logs for the selected section."""
    query = """
        SELECT MIN(date), MAX(date) 
        FROM usage_logs 
        WHERE section_id = %s 
          AND (is_deleted IS NULL OR is_deleted = 0);
    """
    try:
        conn = mariadb.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(query, (section_id,))
        row = cursor.fetchone()
        conn.close()

        if row and row[0] and row[1]:
            d_min = (
                row[0]
                if isinstance(row[0], datetime)
                else datetime.strptime(str(row[0]), "%Y-%m-%d")
            )
            d_max = (
                row[1]
                if isinstance(row[1], datetime)
                else datetime.strptime(str(row[1]), "%Y-%m-%d")
            )
            return d_min.date(), d_max.date()
    except Exception:
        pass

    return datetime(2023, 8, 1).date(), datetime(2023, 12, 31).date()


def fetch_report_data(section_id, start_date, end_date):
    """Pulls entity codes and aggregated counts from MariaDB."""
    query = """
        SELECT 
            entity, 
            COUNT(*) AS total_count
        FROM usage_logs
        WHERE section_id = %s
          AND date BETWEEN %s AND %s
          AND (is_deleted IS NULL OR is_deleted = 0)
        GROUP BY entity
        ORDER BY entity ASC;
    """

    students = []
    employees = []

    try:
        conn = mariadb.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(query, (section_id, start_date, end_date))
        rows = cursor.fetchall()
        conn.close()

        for entity_code, count in rows:
            if not entity_code:
                continue

            if entity_code.upper() in ["FAC", "STAFF"]:
                employees.append((entity_code, count))
            else:
                students.append((entity_code, count))

    except Exception as e:
        messagebox.showerror("Database Error", f"Query Execution Failed:\n{e}")

    return students, employees


def format_dynamic_date_range(start_str, end_str):
    """Formats date range (e.g., 'August to December 2023')."""
    try:
        d1 = datetime.strptime(start_str, "%Y-%m-%d")
        d2 = datetime.strptime(end_str, "%Y-%m-%d")
        if d1.year == d2.year:
            return f"{d1.strftime('%B')} to {d2.strftime('%B %Y')}"
        return f"{d1.strftime('%B %Y')} to {d2.strftime('%B %Y')}"
    except ValueError:
        return f"{start_str} to {end_str}"


class LibraryReportApp:

    def __init__(self, root):
        self.root = root
        self.root.title("JLDM Library Statistics Report")
        self.root.geometry("520x780")

        # Fetch sections from DB
        self.sections_map = fetch_sections()

        # Cache data for Excel export
        self.cached_students = []
        self.cached_employees = []
        self.cached_section_name = ""
        self.cached_date_text = ""

        # --- Top Action Bar / Controls ---
        filter_card = tb.Labelframe(
            root, text=" Report Controls ", bootstyle="primary", padding=15
        )
        filter_card.pack(padx=15, pady=10, fill=X)

        # Area Select
        tb.Label(
            filter_card, text="Area:", font=("Helvetica", 10, "bold")
        ).grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.combo_section = tb.Combobox(
            filter_card,
            values=list(self.sections_map.keys()),
            state="readonly",
            bootstyle="primary",
            width=25,
        )

        default_section = (
            "JLDM Lib - Main Entrance"
            if "JLDM Lib - Main Entrance" in self.sections_map
            else (
                list(self.sections_map.keys())[0] if self.sections_map else ""
            )
        )
        self.combo_section.set(default_section)
        self.combo_section.grid(
            row=0, column=1, columnspan=3, padx=5, pady=5, sticky=W
        )
        self.combo_section.bind("<<ComboboxSelected>>", self.on_section_change)

        # Date Pickers
        tb.Label(
            filter_card, text="From:", font=("Helvetica", 10, "bold")
        ).grid(row=1, column=0, padx=5, pady=5, sticky=W)
        self.cal_start = DateEntry(
            filter_card, bootstyle="primary", dateformat="%Y-%m-%d", width=12
        )
        self.cal_start.grid(row=1, column=1, padx=5, pady=5)

        tb.Label(filter_card, text="To:", font=("Helvetica", 10, "bold")).grid(
            row=1, column=2, padx=5, pady=5, sticky=W
        )
        self.cal_end = DateEntry(
            filter_card, bootstyle="primary", dateformat="%Y-%m-%d", width=12
        )
        self.cal_end.grid(row=1, column=3, padx=5, pady=5)

        # Buttons Panel
        btn_frame = tb.Frame(filter_card)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=(10, 0), sticky=EW)

        btn_generate = tb.Button(
            btn_frame,
            text="Generate",
            bootstyle="success",
            command=self.update_report,
            width=12,
        )
        btn_generate.pack(side=LEFT, padx=5)

        btn_export = tb.Button(
            btn_frame,
            text="Export to .xlsx",
            bootstyle="info-outline",
            command=self.export_to_excel,
            width=14,
        )
        btn_export.pack(side=LEFT, padx=5)

        # Load date bounds
        self.load_date_bounds()

        # --- Table Display Container ---
        self.table_frame = tb.Frame(root, borderwidth=1, relief="solid")
        self.table_frame.pack(padx=15, pady=10, fill=BOTH, expand=True)

        self.table_frame.columnconfigure(0, weight=1)
        self.table_frame.columnconfigure(1, weight=2)
        self.table_frame.columnconfigure(2, weight=1)

        self.update_report()

    def load_date_bounds(self):
        sec_name = self.combo_section.get()
        sec_id = self.sections_map.get(sec_name, 100)

        min_d, max_d = fetch_date_bounds(sec_id)

        # Set calendar default selected values to the database range
        self.cal_start.entry.delete(0, tk.END)
        self.cal_start.entry.insert(0, min_d.strftime("%Y-%m-%d"))

        self.cal_end.entry.delete(0, tk.END)
        self.cal_end.entry.insert(0, max_d.strftime("%Y-%m-%d"))

    def on_section_change(self, event):
        self.load_date_bounds()
        self.update_report()

    def create_cell(
        self,
        parent,
        row,
        col,
        text,
        bg_color="#FFFF00",
        font_weight="normal",
        align="w",
    ):
        lbl = tk.Label(
            parent,
            text=text,
            bg=bg_color,
            fg="#000000",
            font=("Segoe UI", 10, font_weight),
            bd=1,
            relief="solid",
            anchor=align,
            padx=8,
            pady=4,
        )
        lbl.grid(row=row, column=col, sticky="nsew")
        return lbl

    def get_date_str(self, date_widget):
        """Helper to reliably extract YYYY-MM-DD from the DateEntry widget."""
        val = date_widget.entry.get().strip()
        try:
            dt = datetime.strptime(val, "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            return datetime.now().strftime("%Y-%m-%d")

    def update_report(self):
        sec_name = self.combo_section.get()
        sec_id = self.sections_map.get(sec_name, 100)

        start_date = self.get_date_str(self.cal_start)
        end_date = self.get_date_str(self.cal_end)

        for widget in self.table_frame.winfo_children():
            widget.destroy()

        students, employees = fetch_report_data(sec_id, start_date, end_date)

        # Cache variables for Excel exporter
        self.cached_students = students
        self.cached_employees = employees
        self.cached_section_name = sec_name
        self.cached_date_text = format_dynamic_date_range(start_date, end_date)

        curr_row = 0

        # Header Block
        self.create_cell(
            self.table_frame, curr_row, 0, "Area:", font_weight="bold"
        )
        self.create_cell(
            self.table_frame, curr_row, 1, sec_name, font_weight="bold"
        )
        self.create_cell(self.table_frame, curr_row, 2, "", bg_color="#FFFF00")
        curr_row += 1

        self.create_cell(
            self.table_frame, curr_row, 0, "Date:", font_weight="bold"
        )
        self.create_cell(
            self.table_frame,
            curr_row,
            1,
            self.cached_date_text,
            font_weight="bold",
        )
        self.create_cell(self.table_frame, curr_row, 2, "", bg_color="#FFFF00")
        curr_row += 1

        # Spacer
        self.create_cell(self.table_frame, curr_row, 0, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 1, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 2, "", bg_color="#FFFFFF")
        curr_row += 1

        # Student Section
        student_total = 0
        for code, count in students:
            self.create_cell(self.table_frame, curr_row, 0, code)
            self.create_cell(self.table_frame, curr_row, 1, "")
            self.create_cell(
                self.table_frame, curr_row, 2, f"{count:,}", align="e"
            )
            student_total += count
            curr_row += 1

        # Subtotal Row
        self.create_cell(
            self.table_frame,
            curr_row,
            0,
            "TOTAL",
            bg_color="#70AD47",
            font_weight="bold",
        )
        self.create_cell(
            self.table_frame, curr_row, 1, "", bg_color="#70AD47"
        )
        self.create_cell(
            self.table_frame,
            curr_row,
            2,
            f"{student_total:,}",
            bg_color="#70AD47",
            font_weight="bold",
            align="e",
        )
        curr_row += 1

        # Spacer
        self.create_cell(self.table_frame, curr_row, 0, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 1, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 2, "", bg_color="#FFFFFF")
        curr_row += 1

        # Faculty / Staff Section
        employee_total = 0
        for code, count in employees:
            self.create_cell(self.table_frame, curr_row, 0, code)
            self.create_cell(self.table_frame, curr_row, 1, "")
            self.create_cell(
                self.table_frame, curr_row, 2, f"{count:,}", align="e"
            )
            employee_total += count
            curr_row += 1

        # Subtotal Row
        self.create_cell(
            self.table_frame,
            curr_row,
            0,
            "TOTAL",
            bg_color="#70AD47",
            font_weight="bold",
        )
        self.create_cell(
            self.table_frame, curr_row, 1, "", bg_color="#70AD47"
        )
        self.create_cell(
            self.table_frame,
            curr_row,
            2,
            f"{employee_total:,}",
            bg_color="#70AD47",
            font_weight="bold",
            align="e",
        )
        curr_row += 1

        # Spacer
        self.create_cell(self.table_frame, curr_row, 0, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 1, "", bg_color="#FFFFFF")
        self.create_cell(self.table_frame, curr_row, 2, "", bg_color="#FFFFFF")
        curr_row += 1

        # Grand Total
        grand_total = student_total + employee_total
        self.create_cell(
            self.table_frame,
            curr_row,
            0,
            "G.TOTAL",
            bg_color="#FF0000",
            font_weight="bold",
        )
        self.create_cell(
            self.table_frame, curr_row, 1, "", bg_color="#FF0000"
        )
        self.create_cell(
            self.table_frame,
            curr_row,
            2,
            f"{grand_total:,}",
            bg_color="#FF0000",
            font_weight="bold",
            align="e",
        )

    def export_to_excel(self):
        """Exports the current view into a styled .xlsx Excel spreadsheet."""
        if not self.cached_students and not self.cached_employees:
            messagebox.showwarning(
                "Warning", "No report data available to export."
            )
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Library Statistics Report As",
            initialfile=f"Library_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Statistics Report"

            # Use patternType and fgColor for maximum compatibility
            fill_yellow = PatternFill(patternType="solid", fgColor="FFFF00")
            fill_green = PatternFill(patternType="solid", fgColor="70AD47")
            fill_red = PatternFill(patternType="solid", fgColor="FF0000")

            font_bold = Font(name="Arial", size=10, bold=True)
            font_normal = Font(name="Arial", size=10)

            # Use border_style instead of style
            thin_side = Side(border_style="thin", color="000000")
            thin_border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )

            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            def style_cell(
                    cell,
                    value,
                    fill=fill_yellow,
                    font=font_normal,
                    align=align_left,
            ):
                cell.value = value
                cell.fill = fill
                cell.font = font
                cell.alignment = align
                cell.border = thin_border

            current_r = 1

            # Header
            ws.row_dimensions[current_r].height = 20
            style_cell(
                ws.cell(row=current_r, column=1), "Area:", font=font_bold
            )
            style_cell(
                ws.cell(row=current_r, column=2),
                self.cached_section_name,
                font=font_bold,
            )
            style_cell(ws.cell(row=current_r, column=3), "")
            current_r += 1

            ws.row_dimensions[current_r].height = 20
            style_cell(
                ws.cell(row=current_r, column=1), "Date:", font=font_bold
            )
            style_cell(
                ws.cell(row=current_r, column=2),
                self.cached_date_text,
                font=font_bold,
            )
            style_cell(ws.cell(row=current_r, column=3), "")
            current_r += 2

            # Students
            student_tot = 0
            for code, count in self.cached_students:
                ws.row_dimensions[current_r].height = 20
                style_cell(ws.cell(row=current_r, column=1), code)
                style_cell(ws.cell(row=current_r, column=2), "")
                style_cell(
                    ws.cell(row=current_r, column=3), count, align=align_right
                )
                student_tot += count
                current_r += 1

            # Subtotal
            ws.row_dimensions[current_r].height = 20
            style_cell(
                ws.cell(row=current_r, column=1),
                "TOTAL",
                fill=fill_green,
                font=font_bold,
            )
            style_cell(ws.cell(row=current_r, column=2), "", fill=fill_green)
            style_cell(
                ws.cell(row=current_r, column=3),
                student_tot,
                fill=fill_green,
                font=font_bold,
                align=align_right,
            )
            current_r += 2

            # Employees
            employee_tot = 0
            for code, count in self.cached_employees:
                ws.row_dimensions[current_r].height = 20
                style_cell(ws.cell(row=current_r, column=1), code)
                style_cell(ws.cell(row=current_r, column=2), "")
                style_cell(
                    ws.cell(row=current_r, column=3), count, align=align_right
                )
                employee_tot += count
                current_r += 1

            # Subtotal
            ws.row_dimensions[current_r].height = 20
            style_cell(
                ws.cell(row=current_r, column=1),
                "TOTAL",
                fill=fill_green,
                font=font_bold,
            )
            style_cell(ws.cell(row=current_r, column=2), "", fill=fill_green)
            style_cell(
                ws.cell(row=current_r, column=3),
                employee_tot,
                fill=fill_green,
                font=font_bold,
                align=align_right,
            )
            current_r += 2

            # Grand Total
            grand_tot = student_tot + employee_tot
            ws.row_dimensions[current_r].height = 20
            style_cell(
                ws.cell(row=current_r, column=1),
                "G.TOTAL",
                fill=fill_red,
                font=font_bold,
            )
            style_cell(ws.cell(row=current_r, column=2), "", fill=fill_red)
            style_cell(
                ws.cell(row=current_r, column=3),
                grand_tot,
                fill=fill_red,
                font=font_bold,
                align=align_right,
            )

            # Auto column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 15

            wb.save(file_path)
            messagebox.showinfo(
                "Export Success",
                f"Report successfully exported to:\n{file_path}",
            )

        except Exception as e:
            messagebox.showerror(
                "Export Error", f"Failed to save Excel file:\n{e}"
            )


if __name__ == "__main__":
    root = tb.Window(themename="flatly")
    app = LibraryReportApp(root)
    root.mainloop()